from openpyxl import load_workbook

# Define the file path
file_path = r"D:\Needs&Wants.xlsx"
from logger_config import get_logger
logger = get_logger("Get_Product_Links.py")
# Load the existing workbook
def Product_link_from_file(file_path):

    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active  # Get the active worksheet

        # Target cell to read (e.g., A1)
        cell = sheet["F4"]

        # Get the cell value (display text)
        cell_value = cell.value
        if cell_value is None:
            raise Exception
        # Get the hyperlink (if any)
        cell_hyperlink = cell.hyperlink.target if cell.hyperlink else None

    # Output the results
        print(f"Cell Value: {cell_value}")
        print(f"Hyperlink: {cell_hyperlink}")
        return cell_value
    except Exception as e:
        print("To Proceed further links are not available, please check the sheet and update it properly.")




def get_column_data_by_header(file_path, header_name):
    """
    Finds a column by its header name and retrieves all cell data, including coordinates and hyperlinks,
    starting from the first row after the header.

    :param file_path: Path to the Excel file
    :param header_name: Header name to search for
    :return: List of dictionaries with cell data (coordinate, value, and hyperlink)
    """
    # Load the workbook
    workbook = load_workbook(file_path)
    sheet = workbook.active  # Get the active worksheet

    # Find the column with the desired header
    column_index = None
    for col_cells in sheet.iter_cols():  # Iterate through all columns
        for row in range(1,3):
            cell = col_cells[row-1]
            if cell.value == header_name:
                # print(cell.value,"->",cell.column)
                column_index = cell.column
                break

        if column_index is not None:
            break  # Exit the outer loop once header is found

    if column_index is None:
        logger.info(f"Header '{header_name}' not found!")
        return []

    column_data = []

    for col_cell in sheet.iter_cols(min_col=column_index, max_col=column_index, min_row=3, max_row=sheet.max_row):
        for cell in col_cell:
            # print(cell.coordinate)
            cell_data = {
                "coordinate": cell.coordinate,
                "value": cell.value,
            }
            column_data.append(cell_data)
    coordinates_with_values = {}
    for cell_data in column_data:
        coordinates_with_values[cell_data['coordinate']]=cell_data['value']

    # print(coordinates_with_values)
    print("=>$ Received Product link from File")

    return coordinates_with_values
# get_column_data_by_header(file_path,"Links")



