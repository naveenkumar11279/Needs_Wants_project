

from openpyxl import load_workbook

from logger_config import get_logger

logger = get_logger("Load_to_File.py")
# pd = {'F3': 'https://www.amazon.in/dp/B08BC2V71W/?coliid=I13W05AKFJVOGM&colid=1ICFPUUIY9MU9&ref_=list_c_wl_lv_ov_lig_dp_it&th=1', 'F4': 'https://www.amazon.in/dp/B01FP2AB4W/?coliid=IJ4ZB8DKZDVCW&colid=1ICFPUUIY9MU9&ref_=list_c_wl_lv_ov_lig_dp_it&th=1', 'F5': None}

# Define the file path
file_path = r"D:\Needs&Wants.xlsx"

def Price_writing_excel(price_float,ind):
# Load the existing workbook
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active  # Get the active worksheet
        target_cell = sheet[ind]


        # Get the current value in the cell
        current_value = target_cell.value

        # Prepare the updated value
        if current_value is None:
            # If the cell is empty, start with the new price
            updated_value = float(price_float)
            print(f"Cell is empty and i am loading this value : {updated_value}")
            logger.info(f"Cell is empty and i am loading this value : {updated_value}")
        elif current_value == price_float:
            print("Product Price Already is up-to-date, Hence we are not updating in the sheet")
            logger.info("Product Price Already is up-to-date, Hence we are not updating in the sheet")
            raise Exception
        else:
            # Append the new price to the existing value
            updated_value = f"{current_value},{price_float}"
            logger.info("yesterday price, Today price")

        # Write the updated value back to the cell
        target_cell.value = updated_value
        workbook.save(file_path)
        print(f"Value {price_float} written to Excel successfully!")
        logger.info(f"Value {price_float} written to Excel successfully!")
    except Exception as e:
        print(f"We are Breaking up the connection : {e}")
        logger.info(f"We are Breaking up the connection : {e}")




# Price_writing_excel(255)
# print(pd)
def write_to_DatawriteCell(pd,price_float):

    workbook = load_workbook(file_path)
    sheet = workbook.active
    header_name="Data_write"
    column_index = None
    for col_cells in sheet.iter_cols():  # Iterate through all columns
        for row in range(1,3):
            cell = col_cells[row-1]
            if cell.value == header_name:
                # print(cell.value,"->",cell.column)
                column_index = cell.column
                break

        if column_index is not None:
            break  # Exit the outer loop once header is found

    if column_index is None:
        print(f"Header '{header_name}' not found!")
        return []

    for col_write in sheet.iter_cols(column_index,column_index,3,sheet.max_row):

        for col_w in col_write:
            for col,val in pd.items():
                if int(''.join(filter(str.isdigit, col))) == int(''.join(filter(str.isdigit, col_w.coordinate))):
                    target_cell = sheet[col_w.coordinate]
                    current_value = target_cell.value
                    if current_value is None:
                        updated_value = float(price_float)
                    elif current_value == price_float:
                        print("Product Price Already is up-to-date, Hence we are not updating in the sheet")
                        raise Exception
                    else:
                        # Append the new price to the existing value
                        updated_value = f"{current_value},{price_float}"
                    target_cell.value=updated_value
                    print(f"Value {price_float} written to Excel successfully!")
                    break


    workbook.save(file_path)

# write_to_DatawriteCell(pd,300)